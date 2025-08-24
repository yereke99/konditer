#!/usr/bin/env python3.8
"""
Migration script to transfer data from Excel to SQLite database
Creates konditer.db database and migrates user data from just_users 5.xlsx
No pandas dependency - uses openpyxl directly
"""

import sqlite3
import os
import sys
from datetime import datetime

def create_database_and_table():
    """Create SQLite database and the 'just' table"""
    try:
        # Create or connect to database
        conn = sqlite3.connect('konditer.db')
        cursor = conn.cursor()
        
        # Create the table
        table_sql = """CREATE TABLE IF NOT EXISTS just(
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            id_user BIGINT,
                            userName VARCHAR(255),
                            dataRegistred VARCHAR(255)
                        )"""
        
        cursor.execute(table_sql)
        conn.commit()
        print("✓ Database 'konditer.db' created successfully")
        print("✓ Table 'just' created successfully")
        
        return conn
        
    except sqlite3.Error as e:
        print(f"❌ Database error: {e}")
        sys.exit(1)

def read_excel_file(filename):
    """Read Excel file using openpyxl and return data"""
    try:
        from openpyxl import load_workbook
        
        if not os.path.exists(filename):
            print(f"❌ Excel file '{filename}' not found!")
            sys.exit(1)
        
        # Load workbook
        wb = load_workbook(filename, read_only=True, data_only=True)
        ws = wb.active
        
        # Get all data as list of rows
        data = []
        headers = []
        
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_num == 1:
                # First row as headers
                headers = [str(cell).strip().lower() if cell else f'col_{i}' for i, cell in enumerate(row)]
                print(f"✓ Headers found: {headers}")
            else:
                # Data rows
                if any(cell is not None for cell in row):  # Skip empty rows
                    data.append(row)
        
        wb.close()
        
        print(f"✓ Excel file loaded successfully")
        print(f"✓ Found {len(data)} records")
        
        return headers, data
        
    except ImportError:
        print("❌ openpyxl not found! Installing...")
        try:
            import subprocess
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
            print("✓ openpyxl installed. Please run the script again.")
            sys.exit(0)
        except:
            print("❌ Could not install openpyxl. Please install manually:")
            print("pip install openpyxl")
            sys.exit(1)
            
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        sys.exit(1)

def map_excel_data(headers, data):
    """Map Excel data to database columns"""
    # Find column indices for our target fields
    id_user_col = None
    username_col = None
    date_col = None
    
    # Try to find id_user column
    id_user_candidates = ['id_user', 'user_id', 'userid', 'id', 'user id']
    for i, header in enumerate(headers):
        if any(candidate in header for candidate in id_user_candidates):
            id_user_col = i
            print(f"✓ Found id_user column at index {i}: '{headers[i]}'")
            break
    
    # Try to find userName column
    username_candidates = ['username', 'user_name', 'name', 'user name', 'login']
    for i, header in enumerate(headers):
        if any(candidate in header for candidate in username_candidates):
            username_col = i
            print(f"✓ Found userName column at index {i}: '{headers[i]}'")
            break
    
    # Try to find registration date column
    date_candidates = ['dataregistred', 'date_registered', 'registration_date', 'created_at', 'date']
    for i, header in enumerate(headers):
        if any(candidate in header for candidate in date_candidates):
            date_col = i
            print(f"✓ Found date column at index {i}: '{headers[i]}'")
            break
    
    # If not found, use defaults
    if id_user_col is None and len(headers) >= 1:
        id_user_col = 0
        print(f"⚠️  Using first column as id_user: '{headers[0]}'")
    
    if username_col is None and len(headers) >= 2:
        username_col = 1
        print(f"⚠️  Using second column as userName: '{headers[1]}'")
        
    if date_col is None and len(headers) >= 3:
        date_col = 2
        print(f"⚠️  Using third column as date: '{headers[2]}'")
    
    # Process data
    processed_data = []
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    for i, row in enumerate(data):
        try:
            # Extract id_user
            if id_user_col is not None and len(row) > id_user_col:
                id_user = row[id_user_col]
                if id_user is None:
                    id_user = i + 1  # Use row number as fallback
                else:
                    try:
                        id_user = int(float(str(id_user)))
                    except:
                        id_user = i + 1
            else:
                id_user = i + 1
            
            # Extract userName
            if username_col is not None and len(row) > username_col:
                username = row[username_col]
                if username is None:
                    username = f'User_{i+1}'
                else:
                    username = str(username).strip()
            else:
                username = f'User_{i+1}'
            
            # Extract date
            if date_col is not None and len(row) > date_col:
                reg_date = row[date_col]
                if reg_date is None:
                    reg_date = current_date
                else:
                    reg_date = str(reg_date)
            else:
                reg_date = current_date
            
            processed_data.append((id_user, username, reg_date))
            
        except Exception as e:
            print(f"⚠️  Error processing row {i+1}: {e}")
            # Add with defaults
            processed_data.append((i + 1, f'User_{i+1}', current_date))
    
    return processed_data

def migrate_data(conn, data):
    """Insert data into SQLite database"""
    try:
        cursor = conn.cursor()
        
        # Clear existing data (optional - uncomment if needed)
        # cursor.execute("DELETE FROM just")
        
        # Insert data in batches for better performance
        batch_size = 1000
        total_records = len(data)
        
        insert_sql = """INSERT INTO just (id_user, userName, dataRegistred) 
                       VALUES (?, ?, ?)"""
        
        successful_inserts = 0
        
        for i in range(0, total_records, batch_size):
            batch = data[i:i + batch_size]
            try:
                cursor.executemany(insert_sql, batch)
                conn.commit()
                successful_inserts += len(batch)
                print(f"✓ Inserted batch {i//batch_size + 1}: {len(batch)} records")
            except Exception as e:
                print(f"❌ Error inserting batch {i//batch_size + 1}: {e}")
                # Try inserting records one by one
                for record in batch:
                    try:
                        cursor.execute(insert_sql, record)
                        conn.commit()
                        successful_inserts += 1
                    except Exception as record_error:
                        print(f"❌ Error inserting record {record}: {record_error}")
        
        print(f"✓ Migration completed! Successfully inserted: {successful_inserts} records")
        
        # Verify the data
        cursor.execute("SELECT COUNT(*) FROM just")
        count = cursor.fetchone()[0]
        print(f"✓ Verification: {count} records in database")
        
    except Exception as e:
        print(f"❌ Error during migration: {e}")
        conn.rollback()
        sys.exit(1)

def show_sample_data(conn):
    """Display sample of migrated data"""
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM just LIMIT 5")
        rows = cursor.fetchall()
        
        print("\n📋 Sample data from database:")
        print("ID | ID_USER | USERNAME | DATE_REGISTERED")
        print("-" * 50)
        for row in rows:
            print(f"{row[0]} | {row[1]} | {row[2][:30]}{'...' if len(str(row[2])) > 30 else ''} | {row[3]}")
        
    except Exception as e:
        print(f"❌ Error displaying sample data: {e}")

def main():
    """Main migration function"""
    print("🚀 Starting migration from Excel to SQLite...")
    print("=" * 50)
    
    # Create database and table
    conn = create_database_and_table()
    
    # Read Excel file
    excel_filename = './just_users.xlsx'
    headers, raw_data = read_excel_file(excel_filename)
    
    # Map and process data
    processed_data = map_excel_data(headers, raw_data)
    
    # Migrate data
    migrate_data(conn, processed_data)
    
    # Show sample data
    show_sample_data(conn)
    
    # Close connection
    conn.close()
    
    print("\n✅ Migration completed successfully!")
    print(f"Database saved as: konditer.db")

if __name__ == "__main__":
    main()